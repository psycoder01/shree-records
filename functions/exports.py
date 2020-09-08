from openpyxl import load_workbook
from datetime import datetime
from functions.home import getStoreFile

fileLocation = getStoreFile()

def addExports(self):
    name = self.ui.inputNameExports.text()
    product = self.ui.inputProductExports.text()
    total = self.ui.inputTotalExports.text()
    pno = self.ui.inputPnoExports.text()

    data = []
    data.append(name)
    data.append(product)
    data.append(int(total))
    data.append(int(pno))
    data.append(datetime.today().strftime('%Y-%m-%d;%H-%M'))

    wb = load_workbook(fileLocation)
    ws = wb.get_sheet_by_name('exports')
    rowLen = len(ws['A'])
    for col in range(1, 7):
        cell = ws.cell(row=rowLen+1, column=col)
        if(col == 1):
            cell.value = rowLen
            continue
        cell.value = data[col-2]
    wb.save(fileLocation)

    clearInputsExports(self)

def clearInputsExports(self):
    self.ui.inputNameExports.setText("")
    self.ui.inputProductExports.setText("")
    self.ui.inputTotalExports.setText("")
    self.ui.inputPnoExports.setText("")
