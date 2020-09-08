from openpyxl import load_workbook
from datetime import datetime
from functions.home import getStoreFile

fileLocation = getStoreFile()

def addImports(self):
    name = self.ui.inputNameImports.text()
    product = self.ui.inputProductImports.text()
    total = self.ui.inputTotalImports.text()
    pno = self.ui.inputPnoImports.text()

    data = []
    data.append(name)
    data.append(product)
    data.append(int(total))
    data.append(int(pno))
    data.append(datetime.today().strftime('%Y-%m-%d;%H-%M'))

    wb = load_workbook(fileLocation)
    ws = wb.get_sheet_by_name('imports')
    rowLen = len(ws['A'])
    for col in range(1, 7):
        cell = ws.cell(row=rowLen+1, column=col)
        if(col == 1):
            cell.value = rowLen
            continue
        cell.value = data[col-2]
    wb.save(fileLocation)

    clearInputsImports(self)

def clearInputsImports(self):
    self.ui.inputNameImports.setText("")
    self.ui.inputProductImports.setText("")
    self.ui.inputTotalImports.setText("")
    self.ui.inputPnoImports.setText("")
