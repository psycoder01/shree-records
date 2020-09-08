from openpyxl import load_workbook
import pandas as pd
from PySide2.QtWidgets import QTableWidgetItem
from PySide2.QtGui import QFont
from functions.home import getStoreFile

fileName = getStoreFile()

def showDataProducts(self):
    dataFrame = pd.read_excel(fileName, 0)
    rowLength = len(dataFrame.index)
    # Setting max row count in table
    self.ui.tableProducts.setRowCount(rowLength)
    # Showing Data in table
    for row in range(0, rowLength):
        self.ui.tableProducts.setRowHeight(row, 50)
        for col in range(0, 3):
            self.ui.tableProducts.setFont(QFont('Arial', 14))
            self.ui.tableProducts.setItem(
                row, col, QTableWidgetItem(str(dataFrame.values[row, col+1])))


def addProduct(self):
    data = []
    product = self.ui.inputProductProducts.text()
    price = self.ui.inputPriceProducts.text()
    available = self.ui.inputAvlProducts.text()
    data.append(product)
    data.append(price)
    data.append(available)

    wb = load_workbook(fileName)
    ws = wb.get_sheet_by_name('products')
    rowLen = len(ws['A'])
    for col in range(1, 5):
        cell = ws.cell(row=rowLen+1, column=col)
        if(col == 1):
            cell.value = rowLen
            continue
        cell.value = data[col-2]
    wb.save(fileName)
    resetInputs(self)
    showDataProducts(self)


def indexCol(ws, excelIndex, total):
    for row in range(int(excelIndex)+2, total+1):
        cell = ws.cell(row, 1)
        cell.value = row-1


def deleteProduct(self):
    deleteIndex = self.ui.tableProducts.currentRow()
    wb = load_workbook(fileName)
    ws = wb.get_sheet_by_name('products')
    ws.delete_rows(int(deleteIndex)+2)
    indexCol(ws, deleteIndex, len(ws['A']))
    wb.save(fileName)
    showDataProducts(self)


def resetInputs(self):
    self.ui.inputProductProducts.setText("")
    self.ui.inputPriceProducts.setText("")
    self.ui.inputAvlProducts.setText("")
    # updateProduct(self)


def updateProduct(self):
    table = self.ui.tableProducts
    wb = load_workbook(fileName)
    ws = wb.get_sheet_by_name('products')
    rowLen = len(ws['A'])
    for row in range(0, rowLen-1):
        for col in range(0, 3):
            value = table.item(row, col).text()
            cell = ws.cell(row+2, col+2)
            cell.value = value
    wb.save(fileName)
    showDataProducts(self)
