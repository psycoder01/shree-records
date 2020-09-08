import pandas as pd
from openpyxl import load_workbook
from PySide2.QtWidgets import QTableWidgetItem
from PySide2.QtCore import QDate
from PySide2.QtGui import QFont
from functions.home import getStoreFile

fileName =  getStoreFile()

def setDate(self):
    date = QDate.currentDate()
    self.ui.inputDate.setDate(date)


def clearInputs(self):
    self.ui.inputProductSearch.setText("")
    self.ui.inputNameSearch.setText("")
    date = QDate(2020,1,1)
    self.ui.inputDate.setDate(date)

def search(self):
    productName = self.ui.inputProductSearch.text()
    userName = self.ui.inputNameSearch.text()
    option = self.ui.optionsSearch.currentText()
    date = self.ui.inputDate.text()
    if(date == '2020-01-01'):
        date = "" 
   
    if(option == "Imports"):
        df = pd.read_excel(fileName, 1)
    if(option == "Exports"):
        df = pd.read_excel(fileName, 2)
    search = df[(df['Product'].str.contains(productName, case=False))
                & (df['Name'].str.contains(userName, case=False)) & (df['Date'].str.contains(date))]

    rowLength = len(search.index)
    self.ui.tableResultSearch.setRowCount(rowLength)
    self.ui.tableResultSearch.setColumnCount(6)
    # Column Widths
    self.ui.tableResultSearch.setColumnWidth(0, 10)
    self.ui.tableResultSearch.setColumnWidth(1, 180)
    self.ui.tableResultSearch.setColumnWidth(2, 150)
    self.ui.tableResultSearch.setColumnWidth(3, 130)
    self.ui.tableResultSearch.setColumnWidth(4, 150)
    self.ui.tableResultSearch.setColumnWidth(5, 180)

    self.ui.tableResultSearch.verticalHeader().setVisible(False)
    self.ui.tableResultSearch.setHorizontalHeaderLabels(search.head())
    for row in range(0, rowLength):
        self.ui.tableResultSearch.setRowHeight(row, 25)
        for col in range(0, 6):
            self.ui.tableResultSearch.setFont(QFont('Arial', 14))
            self.ui.tableResultSearch.setItem(
                row, col, QTableWidgetItem(str(search.values[row, col])))


def indexCol(ws, excelIndex, total):
    for row in range(int(excelIndex)+1, total+1):
        cell = ws.cell(row, 1)
        cell.value = row-1


def deleteRow(self):
    deleteIndex = self.ui.tableResultSearch.currentRow()
    excelIndex = self.ui.tableResultSearch.item(deleteIndex, 0).text()
    option = self.ui.optionsSearch.currentText()
    wb = load_workbook(fileName)
    ws = wb.get_sheet_by_name(str(option).lower())
    ws.delete_rows(int(excelIndex)+1)
    indexCol(ws, excelIndex, len(ws['A']))
    wb.save(fileName)
    search(self)
