import pandas as pd
from PySide2.QtWidgets import QTableWidgetItem
from PySide2.QtGui import QFont
from datetime import date
import os
from openpyxl import Workbook
from openpyxl.worksheet.table import Table



def validFileName():
    today = str(date.today().strftime('%Y-%m'))
    checkFileName = "data-" + today + ".xlsx"
    return checkFileName

def getStoreFile():
    return 'store/' + validFileName()

fileLocation = getStoreFile()

def showData(self):
    dataFrame = pd.read_excel(fileLocation, 0)
    rowLength = len(dataFrame.index)
    # Setting max row count in table
    self.ui.table_pageHome.setRowCount(rowLength + 1)
    # Showing Data in table
    for row in range(0, rowLength):
        self.ui.table_pageHome.setRowHeight(row, 60)
        for col in range(0, 3):
            self.ui.table_pageHome.setFont(QFont('Arial', 15))
            self.ui.table_pageHome.setItem(
                row, col, QTableWidgetItem(str(dataFrame.values[row,
                                                                col + 1])))


def checkFile():
    validFileExist = os.path.isfile(fileLocation)
    if (validFileExist):
        print('file already exists')
        return
    else:
        pheaders = ["S.N", "Product", "Rate", "Available"]
        iheaders = ["S.N", "Name", "Product", "Total", "Phone No", "Date"]
        wb = Workbook()
        ws = wb.active
        ws_products = wb.create_sheet(title="products", index=0)
        ws_imports = wb.create_sheet(title="imports", index=1)
        ws_exports = wb.create_sheet(title="exports", index=2)
        wb.remove_sheet(ws)
        ws_products.append(pheaders)
        ws_imports.append(iheaders)
        ws_exports.append(iheaders)
        wb.save(fileLocation)
        print("file saved")

